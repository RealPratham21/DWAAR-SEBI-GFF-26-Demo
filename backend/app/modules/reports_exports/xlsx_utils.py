"""XLSX helpers for Reports & Export (G7)."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def _autosize_columns(sheet: Worksheet, *, max_width: int = 48) -> None:
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(value), max_width))
        sheet.column_dimensions[get_column_letter(col_idx)].width = max(length + 2, 10)


def write_sheet(
    workbook: Workbook,
    title: str,
    headers: list[str],
    rows: Iterable[list[Any]],
) -> None:
    sheet = workbook.create_sheet(title=title[:31])
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _autosize_columns(sheet)


def workbook_bytes(workbook: Workbook) -> bytes:
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        workbook.remove(workbook["Sheet"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
