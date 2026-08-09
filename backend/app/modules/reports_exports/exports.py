"""Register and workbook exports (G7)."""

from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import Workbook

from app.modules.data_room.constants import DOC_STATUS_LABELS, REQUIREMENT_STATUS_LABELS
from app.modules.data_room.labels import workstream_label
from app.modules.facts_evidence.aggregator import aggregate_facts_context
from app.modules.facts_evidence.constants import SUPPORT_STATE_LABELS, SUPPORT_TYPE_LABELS
from app.modules.facts_evidence.service import _usage_for_fact
from app.modules.issues_gaps.labels import chapter_labels
from app.modules.reports_exports.context import ReportsExportContext, facts_with_usage
from app.modules.reports_exports.filenames import dated_suffix, issuer_prefix
from app.modules.reports_exports.xlsx_utils import write_sheet, workbook_bytes


def _yes_no(value: bool) -> str:
    return "Yes" if value else "No"


def export_issues_xlsx(ctx: ReportsExportContext) -> tuple[bytes, str]:
    wb = Workbook()
    headers = [
        "Severity",
        "Status",
        "Category",
        "Issue",
        "Description",
        "Workstream",
        "Section",
        "Record",
        "Why it matters",
        "Suggested action",
        "Professional review required",
        "Affected DRHP chapters",
        "Acknowledged",
        "User note",
    ]
    rows = []
    for issue in ctx.issues:
        rows.append(
            [
                issue.severity,
                issue.lifecycle_state,
                issue.category,
                issue.title,
                issue.description,
                issue.workstream_label,
                issue.section_label,
                issue.record_label,
                issue.why_it_matters,
                issue.suggested_action,
                _yes_no(issue.professional_review_required),
                ", ".join(chapter_labels(issue.affected_drhp_chapters)),
                _yes_no(issue.acknowledged),
                issue.acknowledgement_note or "",
            ]
        )
    write_sheet(wb, "Issues", headers, rows)
    filename = f"{issuer_prefix(ctx.issuer_name)}_Issues_Gaps_{dated_suffix()}.xlsx"
    return workbook_bytes(wb), filename


def export_issues_csv(ctx: ReportsExportContext) -> tuple[bytes, str]:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "Severity",
            "Status",
            "Category",
            "Issue",
            "Description",
            "Workstream",
            "Section",
            "Record",
            "Why it matters",
            "Suggested action",
            "Professional review required",
            "Affected DRHP chapters",
            "Acknowledged",
            "User note",
        ]
    )
    for issue in ctx.issues:
        writer.writerow(
            [
                issue.severity,
                issue.lifecycle_state,
                issue.category,
                issue.title,
                issue.description,
                issue.workstream_label,
                issue.section_label,
                issue.record_label,
                issue.why_it_matters,
                issue.suggested_action,
                _yes_no(issue.professional_review_required),
                ", ".join(chapter_labels(issue.affected_drhp_chapters)),
                _yes_no(issue.acknowledged),
                issue.acknowledgement_note or "",
            ]
        )
    filename = f"{issuer_prefix(ctx.issuer_name)}_Issues_Gaps_{dated_suffix()}.csv"
    return buffer.getvalue().encode("utf-8-sig"), filename


def export_facts_evidence_xlsx(db, user, ctx: ReportsExportContext) -> tuple[bytes, str]:
    facts_ctx = aggregate_facts_context(db, user)
    fact_rows = facts_with_usage(ctx, facts_ctx)

    wb = Workbook()
    write_sheet(
        wb,
        "Facts",
        [
            "Fact",
            "Value",
            "Workstream",
            "Section",
            "Record",
            "Support type",
            "Support state",
            "Period / as-of",
            "Evidence count",
            "DRHP usage",
            "Affected chapters",
            "Related issues",
        ],
        [
            [
                r["label"],
                r["value"],
                r["workstream"],
                r["section"],
                r["record"],
                SUPPORT_TYPE_LABELS.get(r["supportType"], r["supportType"]),
                SUPPORT_STATE_LABELS.get(r["supportState"], r["supportState"]),
                r["period"],
                r["evidenceCount"],
                r["drhpUsageCount"],
                r["affectedChapters"],
                r["relatedIssues"],
            ]
            for r in fact_rows
        ],
    )

    evidence_rows = []
    for item in ctx.evidence:
        evidence_rows.append(
            [
                item.document_name,
                item.document_category,
                item.version_number,
                item.page_number,
                item.assertion_label or item.extracted_text_preview,
                len(item.supported_fact_ids or []),
                item.drhp_usage_count,
                item.processing_state,
            ]
        )

    write_sheet(
        wb,
        "Evidence",
        [
            "Document",
            "Category",
            "Version",
            "Page",
            "Assertion / evidence",
            "Facts supported",
            "DRHP usage",
            "Processing state",
        ],
        evidence_rows,
    )

    drhp_rows = []
    for fact in ctx.facts:
        for usage in _usage_for_fact(fact, facts_ctx.drhp_usage):
            drhp_rows.append(
                [
                    fact.label,
                    fact.display_value,
                    fact.workstream_label,
                    usage.chapter_label,
                    usage.section_heading,
                    usage.block_id,
                ]
            )
    write_sheet(
        wb,
        "DRHP Usage",
        ["Fact", "Value", "Workstream", "Chapter", "Section", "Block"],
        drhp_rows,
    )

    filename = f"{issuer_prefix(ctx.issuer_name)}_Facts_Evidence_{dated_suffix()}.xlsx"
    return workbook_bytes(wb), filename


def export_data_room_xlsx(ctx: ReportsExportContext) -> tuple[bytes, str]:
    wb = Workbook()
    doc_rows = []
    for doc in ctx.data_room_documents:
        inspection = (doc.metadata or {}).get("inspection") or {}
        doc_rows.append(
            [
                doc.title,
                doc.filename,
                workstream_label(doc.workstream_key),
                doc.category,
                doc.requirement_key or "",
                doc.current_version,
                DOC_STATUS_LABELS.get(doc.status, doc.status),
                doc.processing_capability,
                doc.uploaded_at.isoformat() if doc.uploaded_at else "",
                doc.fact_count,
                doc.evidence_count,
                doc.issue_count,
                doc.drhp_usage_count,
                inspection.get("label") or inspection.get("status") or "",
            ]
        )
    write_sheet(
        wb,
        "Uploaded Documents",
        [
            "Document title",
            "Filename",
            "Workstream",
            "Category",
            "Requirement",
            "Version",
            "Status",
            "Processing capability",
            "Uploaded date",
            "Fact count",
            "Evidence count",
            "Issue count",
            "DRHP usage",
            "Inspection-list state",
        ],
        doc_rows,
    )

    req_rows = []
    for req in ctx.data_room_requirements:
        req_rows.append(
            [
                req.title,
                workstream_label(req.workstream_key),
                req.category,
                req.purpose,
                req.applicability_state,
                REQUIREMENT_STATUS_LABELS.get(req.status, req.status),
                ", ".join(req.matched_document_ids),
                _yes_no(req.professional_confirmation_required),
                len(req.linked_issue_ids),
            ]
        )
    write_sheet(
        wb,
        "Expected Documents",
        [
            "Requirement",
            "Workstream",
            "Category",
            "Purpose",
            "Applicability",
            "Status",
            "Provided documents",
            "Professional confirmation required",
            "Related issues",
        ],
        req_rows,
    )
    filename = f"{issuer_prefix(ctx.issuer_name)}_Data_Room_Index_{dated_suffix()}.xlsx"
    return workbook_bytes(wb), filename


def export_preparation_workbook(db, user, ctx: ReportsExportContext) -> tuple[bytes, str]:
    facts_ctx = aggregate_facts_context(db, user)
    wb = Workbook()

    overview = [
        ["Issuer", ctx.issuer_name],
        ["Generated at", ctx.generated_at.isoformat()],
        ["Workstreams complete", ctx.workstream_totals.get("complete", 0)],
        ["Workstreams in progress", ctx.workstream_totals.get("inProgress", 0)],
        ["Workstreams not started", ctx.workstream_totals.get("notStarted", 0)],
        ["Open issues", ctx.issues_summary.get("totalOpen", 0)],
        ["Blocking issues", ctx.issues_summary.get("blocking", 0)],
        ["High issues", ctx.issues_summary.get("high", 0)],
        ["Professional-review items", ctx.issues_summary.get("professionalReview", 0)],
        ["Canonical facts", ctx.facts_summary.get("canonicalFacts", 0)],
        ["Document-backed facts", ctx.facts_summary.get("documentBacked", 0)],
        ["Structured-input facts", ctx.facts_summary.get("structuredInput", 0)],
        ["Uploaded Data Room documents", ctx.data_room_summary.get("totalDocuments", 0)],
        ["Missing applicable document requirements", ctx.data_room_summary.get("missingRequirements", 0)],
        ["Latest DRHP version", ctx.latest_drhp_version_number or "—"],
        ["DRHP status", ctx.latest_drhp_status_label],
        ["DRHP stale", _yes_no(ctx.drhp_stale)],
    ]
    write_sheet(wb, "Overview", ["Metric", "Value"], overview)

    issues_by_ws = ctx.issues_summary.get("byWorkstream") or {}
    ws_rows = []
    req_by_ws: dict[str, dict[str, int]] = {}
    for req in ctx.data_room_requirements:
        bucket = req_by_ws.setdefault(req.workstream_key, {"provided": 0, "missing": 0})
        if req.applicability_state == "not_applicable":
            continue
        if req.status in {"provided", "partially_provided"}:
            bucket["provided"] += 1
        elif req.status == "not_provided":
            bucket["missing"] += 1

    for row in ctx.workstreams:
        key = row["workstreamKey"]
        ws_rows.append(
            [
                row["workstreamLabel"],
                row["overallStatus"],
                row["sectionsComplete"],
                row["totalSections"],
                row["overallStatus"],
                issues_by_ws.get(key, 0),
                sum(1 for i in ctx.issues if i.workstream_key == key and i.professional_review_required),
                req_by_ws.get(key, {}).get("provided", 0),
                req_by_ws.get(key, {}).get("missing", 0),
                "",
            ]
        )
    write_sheet(
        wb,
        "Workstreams",
        [
            "Workstream",
            "Progress state",
            "Completed sections",
            "Total sections",
            "Assessment / review state",
            "Open issues",
            "Professional-review items",
            "Provided document requirements",
            "Missing document requirements",
            "Affected DRHP chapters",
        ],
        ws_rows,
    )

    issue_rows = []
    for issue in ctx.issues:
        issue_rows.append(
            [
                issue.severity,
                issue.lifecycle_state,
                issue.category,
                issue.title,
                issue.workstream_label,
                issue.section_label,
                issue.suggested_action,
            ]
        )
    write_sheet(
        wb,
        "Issues & Gaps",
        ["Severity", "Status", "Category", "Issue", "Workstream", "Section", "Suggested action"],
        issue_rows,
    )

    fact_rows = facts_with_usage(ctx, facts_ctx)
    write_sheet(
        wb,
        "Facts",
        ["Fact", "Value", "Workstream", "Section", "Support type", "Evidence count", "DRHP usage"],
        [
            [
                r["label"],
                r["value"],
                r["workstream"],
                r["section"],
                r["supportType"],
                r["evidenceCount"],
                r["drhpUsageCount"],
            ]
            for r in fact_rows
        ],
    )

    evidence_rows = []
    for item in ctx.evidence:
        evidence_rows.append(
            [
                item.document_name,
                item.document_category,
                item.version_number,
                item.page_number,
                item.assertion_label or item.extracted_text_preview,
                len(item.supported_fact_ids or []),
            ]
        )
    write_sheet(
        wb,
        "Evidence",
        ["Document", "Category", "Version", "Page", "Assertion / evidence", "Facts supported"],
        evidence_rows,
    )

    doc_rows = []
    for doc in ctx.data_room_documents:
        doc_rows.append(
            [
                doc.title,
                doc.filename,
                workstream_label(doc.workstream_key),
                doc.category,
                doc.status,
                doc.fact_count,
                doc.evidence_count,
            ]
        )
    write_sheet(
        wb,
        "Data Room",
        ["Document", "Filename", "Workstream", "Category", "Status", "Facts", "Evidence"],
        doc_rows,
    )

    req_rows = []
    for req in ctx.data_room_requirements:
        req_rows.append(
            [
                req.title,
                workstream_label(req.workstream_key),
                req.category,
                req.applicability_state,
                req.status,
            ]
        )
    write_sheet(wb, "Document Requirements", ["Requirement", "Workstream", "Category", "Applicability", "Status"], req_rows)

    drhp_rows = []
    for idx, chapter in enumerate(ctx.drhp_chapters, start=1):
        drhp_rows.append(
            [
                idx,
                chapter.get("title") or chapter.get("chapterKey"),
                chapter.get("status"),
                len(chapter.get("warnings") or []),
                "",
                "",
                "",
                _yes_no(ctx.drhp_stale),
            ]
        )
    write_sheet(
        wb,
        "DRHP Status",
        [
            "Order",
            "Chapter",
            "Generation status",
            "Warnings",
            "SourceRef count",
            "EvidenceRef count",
            "Placeholder count",
            "Stale/affected",
        ],
        drhp_rows,
    )

    filename = f"{issuer_prefix(ctx.issuer_name)}_IPO_Preparation_Workbook_{dated_suffix()}.xlsx"
    return workbook_bytes(wb), filename
